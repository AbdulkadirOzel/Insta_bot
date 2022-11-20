import pandas as pd

likelist = open(r"C:\Users\Abdulkadir\Desktop\Insta_bot\likelist.txt", "rt")
commentlist = open(r"C:\Users\Abdulkadir\Desktop\Insta_bot\cmmlist.txt", "rt")

# commentlistteki 1. kişiyi listeden çıkarır.

print(likelist.read().count("ozel.abdulkadirr"))


likelist.close()
commentlist.close()
#comments.close()

#%%
import pandas as pd
import numpy as np
from openpyxl import load_workbook

with open(r"C:\Users\Abdulkadir\Desktop\Insta_bot\cmmlist.txt", "r") as comments:
    commentators = [i for i in comments]
    library = {x:commentators.count(x) for x in commentators}
    isimler = list(library.keys()) # yorum yapanların isimleri
    sayilar = list(library.values()) # yorum sayıları
    #print(isimler)
    #print(sayilar)
    

wb = load_workbook(r"C:\Users\Abdulkadir\Desktop\Insta_bot\example_excel.xlsx")
ws = wb["Sheet1"]
ws["A1"] = "Names" #just a header
ws["B1"] = "Comment_numbers"
ws["C1"] = "Like_numbers"
ws["D1"] = "Scores"

row = 2 #represent the 2 row of the sheet
column = 1 #represent the column "A" of the sheet

for i in isimler:
    ws.cell(row=row, column=column).value = i #getting the current cell, and writing the value of the list
    row += 1 #just setting the current to the next

row = 2 #represent the 2 row of the sheet
column = 2 #represent the column "A" of the sheet

for i in sayilar:
    ws.cell(row=row, column=column).value = i #getting the current cell, and writing the value of the list
    row += 1 #just setting the current to the next

row = 2
column = 3

likelist = open(r"C:\Users\Abdulkadir\Desktop\Insta_bot\likelist.txt", "rt")

likers = [i for i in likelist]

for i in isimler:
    a = likers.count(i)
    ws.cell(row=row, column=column).value = a
    row += 1

wb_sheet = wb.active
wb_sheet['D2'] = '= 2*B2 + C2'  
wb_sheet['D3'] = '= 2*B3 + C3'
wb_sheet['D4'] = '= 2*B4 + C4'
wb_sheet['D5'] = '= 2*B5 + C5'
wb_sheet['D6'] = '= 2*B6 + C6'
wb_sheet['D7'] = '= 2*B7 + C7'
wb_sheet['D8'] = '= 2*B8 + C8'
wb_sheet['D9'] = '= 2*B9 + C9'
wb_sheet['D10'] = '= 2*B10 + C10'
wb_sheet['D11'] = '= 2*B11 + C11'
wb_sheet['D12'] = '= 2*B12 + C12'
wb_sheet['D13'] = '= 2*B13 + C13'
wb_sheet['D14'] = '= 2*B14 + C14'
wb_sheet['D15'] = '= 2*B15 + C15'
wb_sheet['D16'] = '= 2*B16 + C16'
wb_sheet['D17'] = '= 2*B17 + C17'
wb_sheet['D18'] = '= 2*B18 + C18'
wb_sheet['D19'] = '= 2*B19 + C19'
wb_sheet['D20'] = '= 2*B20 + C20'
wb_sheet['D21'] = '= 2*B21 + C21'
wb_sheet['D22'] = '= 2*B22 + C22'
wb_sheet['D23'] = '= 2*B23 + C23'
wb_sheet['D24'] = '= 2*B24 + C24'
wb_sheet['D25'] = '= 2*B25 + C25'
wb_sheet['D26'] = '= 2*B26 + C26'
wb_sheet['D27'] = '= 2*B27 + C27'
wb_sheet['D28'] = '= 2*B28 + C28'
wb_sheet['D29'] = '= 2*B29 + C29'
wb_sheet['D30'] = '= 2*B30 + C30'
wb_sheet['D31'] = '= 2*B31 + C31'
wb_sheet['D32'] = '= 2*B32 + C32'
wb_sheet['D33'] = '= 2*B33 + C33'
wb_sheet['D34'] = '= 2*B34 + C34'


wb.save(r"C:\Users\Abdulkadir\Desktop\Insta_bot\example_excel.xlsx")

comments.close()
likelist.close()
wb.close()
